#!/usr/bin/env python3
"""
Importa a aba "Aulas" de uma planilha XLSX para as tabelas:
- public.lesson_modules
- public.lessons

Uso:
  python scripts/import_lessons_from_xlsx.py
  python scripts/import_lessons_from_xlsx.py --xlsx "C:\\caminho\\arquivo.xlsx"
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


DEFAULT_XLSX = r"C:\Users\WINDOWS\OneDrive\Documentos\Aulas Mentoria CORE Business.xlsx"
DEFAULT_ENV = "supabase/functions/.env"
SHEET_NAME = "Aulas"


@dataclass
class RuntimeConfig:
  supabase_url: str
  service_role_key: str


def normalize_header(value: str | None) -> str:
  if not value:
    return ""
  raw = unicodedata.normalize("NFD", str(value))
  raw = "".join(ch for ch in raw if unicodedata.category(ch) != "Mn")
  raw = re.sub(r"\s+", " ", raw.strip().lower())
  return raw


def read_env_value(file_path: Path, key: str) -> str:
  for line in file_path.read_text(encoding="utf-8").splitlines():
    if line.startswith(f"{key}="):
      return line.split("=", 1)[1].strip()
  raise RuntimeError(f"Chave '{key}' nao encontrada em {file_path}")


def load_runtime(env_path: Path) -> RuntimeConfig:
  return RuntimeConfig(
    supabase_url=read_env_value(env_path, "SUPABASE_URL"),
    service_role_key=read_env_value(env_path, "SUPABASE_SERVICE_ROLE_KEY"),
  )


def request_json(
  cfg: RuntimeConfig,
  method: str,
  path: str,
  payload: Any | None = None,
  prefer: str | None = None,
) -> Any:
  url = f"{cfg.supabase_url}{path}"
  data = None
  headers = {
    "apikey": cfg.service_role_key,
    "Authorization": f"Bearer {cfg.service_role_key}",
  }

  if payload is not None:
    raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    data = raw
    headers["Content-Type"] = "application/json; charset=utf-8"

  if prefer:
    headers["Prefer"] = prefer

  req = urllib.request.Request(url=url, data=data, method=method.upper(), headers=headers)

  try:
    with urllib.request.urlopen(req) as resp:
      body = resp.read().decode("utf-8")
      if not body:
        return None
      return json.loads(body)
  except urllib.error.HTTPError as exc:
    detail = exc.read().decode("utf-8", errors="replace")
    raise RuntimeError(f"REST {method} {path} falhou ({exc.code}): {detail}") from exc


def parse_sheet(xlsx_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
  wb = openpyxl.load_workbook(xlsx_path, data_only=True)
  if SHEET_NAME not in wb.sheetnames:
    raise RuntimeError(f"A planilha nao contem a aba '{SHEET_NAME}'")

  ws = wb[SHEET_NAME]
  if ws.max_row < 2:
    raise RuntimeError("A aba Aulas esta vazia")

  headers = [normalize_header(ws.cell(1, c).value) for c in range(1, ws.max_column + 1)]
  idx_modulo = headers.index("modulo") if "modulo" in headers else -1
  idx_titulo = headers.index("titulo") if "titulo" in headers else -1
  idx_descricao = headers.index("descricao") if "descricao" in headers else -1
  idx_link = headers.index("link loom") if "link loom" in headers else -1

  if idx_modulo < 0 or idx_titulo < 0:
    raise RuntimeError("Cabecalhos esperados nao encontrados (Modulo, Titulo)")

  modules_order: list[str] = []
  lessons_raw: list[dict[str, Any]] = []
  per_module_order: dict[str, int] = {}

  for row in range(2, ws.max_row + 1):
    def cell(index: int) -> str:
      if index < 0:
        return ""
      value = ws.cell(row, index + 1).value
      return str(value).strip() if value is not None else ""

    module_title = cell(idx_modulo)
    lesson_title = cell(idx_titulo)
    description = cell(idx_descricao)
    loom_link = cell(idx_link)

    if not module_title:
      continue

    if module_title not in modules_order:
      modules_order.append(module_title)
      per_module_order[module_title] = 0

    # Linhas com modulo mas sem titulo representam placeholders no XLSX
    if not lesson_title:
      continue

    per_module_order[module_title] += 1
    lessons_raw.append(
      {
        "module_title": module_title,
        "title": lesson_title,
        "description": description or None,
        "loom_id": loom_link,
        "order_index": per_module_order[module_title],
        "is_active": True,
      }
    )

  modules_payload = [
    {
      "title": module_title,
      "description": None,
      "order_index": idx + 1,
      "is_active": True,
    }
    for idx, module_title in enumerate(modules_order)
  ]

  return modules_payload, lessons_raw


def main() -> int:
  parser = argparse.ArgumentParser()
  parser.add_argument("--xlsx", default=DEFAULT_XLSX, help="Caminho da planilha XLSX")
  parser.add_argument("--env", default=DEFAULT_ENV, help="Arquivo .env das functions")
  args = parser.parse_args()

  xlsx_path = Path(args.xlsx)
  env_path = Path(args.env)

  if not xlsx_path.exists():
    raise RuntimeError(f"Arquivo XLSX nao encontrado: {xlsx_path}")
  if not env_path.exists():
    raise RuntimeError(f"Arquivo de ambiente nao encontrado: {env_path}")

  cfg = load_runtime(env_path)
  modules_payload, lessons_raw = parse_sheet(xlsx_path)

  print(f"Planilha lida: {xlsx_path}")
  print(f"Modulos encontrados: {len(modules_payload)}")
  print(f"Aulas validas encontradas: {len(lessons_raw)}")

  # Limpa base atual (lessons antes de modules por FK)
  request_json(
    cfg,
    "DELETE",
    "/rest/v1/lessons?id=neq.00000000-0000-0000-0000-000000000000",
  )
  request_json(
    cfg,
    "DELETE",
    "/rest/v1/lesson_modules?id=neq.00000000-0000-0000-0000-000000000000",
  )

  inserted_modules: list[dict[str, Any]] = []
  for module in modules_payload:
    inserted = request_json(
      cfg,
      "POST",
      "/rest/v1/lesson_modules",
      payload=module,
      prefer="return=representation",
    )
    inserted_modules.extend(inserted or [])

  module_id_by_title = {m["title"]: m["id"] for m in inserted_modules}
  lessons_payload: list[dict[str, Any]] = []
  for lesson in lessons_raw:
    module_id = module_id_by_title.get(lesson["module_title"])
    if not module_id:
      continue
    lessons_payload.append(
      {
        "module_id": module_id,
        "title": lesson["title"],
        "description": lesson["description"],
        # O schema atual usa loom_id NOT NULL; quando nao houver video, envia string vazia.
        "loom_id": lesson["loom_id"] or "",
        "duration": None,
        "order_index": lesson["order_index"],
        "is_active": True,
      }
    )

  # Insercao em lotes
  inserted_lessons = 0
  batch_size = 100
  for i in range(0, len(lessons_payload), batch_size):
    batch = lessons_payload[i : i + batch_size]
    inserted = request_json(
      cfg,
      "POST",
      "/rest/v1/lessons",
      payload=batch,
      prefer="return=representation",
    )
    inserted_lessons += len(inserted or [])

  print("Importacao concluida com sucesso.")
  print(f"Modulos inseridos: {len(inserted_modules)}")
  print(f"Aulas inseridas: {inserted_lessons}")
  return 0


if __name__ == "__main__":
  try:
    raise SystemExit(main())
  except Exception as exc:  # noqa: BLE001
    print(f"ERRO: {exc}", file=sys.stderr)
    raise SystemExit(1)
